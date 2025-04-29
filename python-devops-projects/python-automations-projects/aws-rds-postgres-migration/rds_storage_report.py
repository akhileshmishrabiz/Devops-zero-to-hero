from datetime import datetime, timedelta
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font
import logging
import sys
from typing import List, Tuple, Dict
import argparse
import boto3

rds = boto3.client("rds")
cloudwatch = boto3.client("cloudwatch")

def set_logs(verbose: bool) -> None:
    """
    Configure logging settings.

    :param verbose: Boolean indicating verbosity level.
    """
    logging.basicConfig(
        stream=sys.stdout,
        format="%(asctime)s - %(levelname)s: %(message)s",
        level=logging.DEBUG if verbose else logging.INFO,
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    logging.getLogger("boto3").setLevel(logging.WARNING)
    logging.getLogger("botocore").setLevel(logging.WARNING)
    logging.getLogger("urllib3.connectionpool").setLevel(logging.WARNING)

def get_args() -> argparse.Namespace:
    """
    Parse command line arguments.

    :return: Parsed arguments.
    """
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--output",
        "-o",
        type=str,
        default="rds_storage_report.xlsx",
        help="Output Excel file name",
    )
    return parser.parse_args()

def save_to_excel(
    data: List[Tuple[str, str, str, str, int, float]],
    headers: List[str],
    excel_file: str,
) -> None:
    """
    Save data to an Excel file with the specified headers.
    If the file exists, add the data to the default sheet without overwriting existing data.

    :param data: List of data rows.
    :param headers: List of column headers.
    :param excel_file: Path to the Excel file.
    """
    try:
        # Load existing workbook if it exists
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active  # Use the default sheet
        ws.delete_rows(1, ws.max_row)
    except FileNotFoundError:
        # Create a new workbook if the file does not exist
        wb = openpyxl.Workbook()
        ws = wb.active  # Use the default sheet

    # Write headers to the first row with a larger and bold font
    header_font = Font(bold=True, size=12)  # Slightly larger and bold
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font

    # Write data to subsequent rows
    for row, row_data in enumerate(data, start=2):
        for col, cell_data in enumerate(row_data, start=1):
            ws.cell(row=row, column=col, value=str(cell_data))
            # Calculate the length of the cell data and adjust column width
            column_letter = get_column_letter(col)
            column_width = max(len(str(cell_data)), len(header))
            ws.column_dimensions[column_letter].width = (
                column_width + 2
            )  # Adding extra padding

    # Define table range
    table_range = f"A1:{get_column_letter(len(headers))}{len(data) + 1}"

    # Generate a unique table name using the current timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    table_name = f"Table_{timestamp}"

    # Add table formatting with a higher contrast style
    tab = Table(displayName=table_name, ref=table_range)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Save the workbook
    wb.save(excel_file)

def get_db_freestorage(DBInstanceIdentifier: str) -> float:
    """
    Get the free storage space for a given RDS instance.

    :param DBInstanceIdentifier: The identifier of the RDS instance.
    :return: Free storage space in gigabytes.
    """
    response = cloudwatch.get_metric_data(
        MetricDataQueries=[
            {
                "Id": "fetching_FreeStorageSpace",
                "MetricStat": {
                    "Metric": {
                        "Namespace": "AWS/RDS",
                        "MetricName": "FreeStorageSpace",
                        "Dimensions": [
                            {
                                "Name": "DBInstanceIdentifier",
                                "Value": DBInstanceIdentifier,
                            }
                        ],
                    },
                    "Period": 86400 * 7,
                    "Stat": "Minimum",
                },
            }
        ],
        StartTime=(datetime.now() - timedelta(seconds=86400 * 7)).timestamp(),
        EndTime=datetime.now().timestamp(),
        ScanBy="TimestampDescending",
    )
    try:
        return round(response["MetricDataResults"][0]["Values"][0] / 1024**3, 2)
    except (IndexError, KeyError) as e:
        logging.error(f"Error fetching free storage for {DBInstanceIdentifier}: {e}")
        return 0.0

def conditional(db_instance: Dict[str, str]) -> bool:
    """
    Check if the RDS instance meets certain conditions.

    :param db_instance: The RDS instance dictionary.
    :return: True if conditions are met, False otherwise.
    """
    return (
        db_instance.get("DBInstanceClass", "") != "db.serverless"
        and db_instance.get("DBClusterIdentifier", "") == ""
    )

def format_data_output(db_instances: List[Dict[str, str]]) -> List[Tuple[str, str, str, str, int, float]]:
    """
    Format the RDS instance data for output.

    :param db_instances: List of RDS instance dictionaries.
    :return: Formatted list of tuples containing RDS instance data.
    """
    return [
        (
            db_instance["DBInstanceIdentifier"],
            db_instance["DBInstanceClass"],
            db_instance["Engine"],
            db_instance["EngineVersion"],
            db_instance["AllocatedStorage"],
            get_db_freestorage(db_instance["DBInstanceIdentifier"]),
        )
        for db_instance in db_instances
    ]

def get_rds_instances() -> List[Dict[str, str]]:
    """
    Retrieve the list of RDS instances.

    :return: List of RDS instance dictionaries.
    """
    db_instances = rds.describe_db_instances()["DBInstances"]
    return [db_instance for db_instance in db_instances if conditional(db_instance)]

def run() -> None:
    """
    Main function to execute the script.
    """
    set_logs(verbose=False)
    args = get_args()
    consolidated_rds_formatted_data: List[Tuple[str, str, str, str, int, float]] = []
    
    rds_formatted_data_for_account = format_data_output(get_rds_instances())
    consolidated_rds_formatted_data.extend(rds_formatted_data_for_account)

    header = [
        "DBInstanceIdentifier",
        "DBInstanceClass",
        "DBEngine",
        "Version",
        "AllocatedStorage",
        "FreeStorageSpace",
    ]
    save_to_excel(consolidated_rds_formatted_data, header, args.output)

if __name__ == "__main__":
    run()