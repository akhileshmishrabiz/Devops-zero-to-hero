from datetime import datetime, timedelta
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font
import argparse, boto3

rds, cloudwatch = boto3.client("rds"), boto3.client("cloudwatch")

def get_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", "-o", type=str, default="rds_storage_report.xlsx", help="Output Excel file name")
    return parser.parse_args()

def save_to_excel(data, headers, excel_file):
    try:
        wb, ws = openpyxl.load_workbook(excel_file), openpyxl.load_workbook(excel_file).active
        ws.delete_rows(1, ws.max_row)
    except FileNotFoundError:
        wb, ws = openpyxl.Workbook(), openpyxl.Workbook().active

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12)

    for row, row_data in enumerate(data, start=2):
        for col, cell_data in enumerate(row_data, start=1):
            ws.cell(row=row, column=col, value=str(cell_data))
            ws.column_dimensions[get_column_letter(col)].width = max(len(str(cell_data)), len(header)) + 2

    table_range = f"A1:{get_column_letter(len(headers))}{len(data) + 1}"
    tab = Table(displayName=f"Table_{datetime.now().strftime('%Y%m%d_%H%M%S')}", ref=table_range)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    ws.add_table(tab)
    wb.save(excel_file)

def get_db_freestorage(DBInstanceIdentifier):
    response = cloudwatch.get_metric_data(
        MetricDataQueries=[{
            "Id": "fetching_FreeStorageSpace",
            "MetricStat": {
                "Metric": {
                    "Namespace": "AWS/RDS",
                    "MetricName": "FreeStorageSpace",
                    "Dimensions": [{"Name": "DBInstanceIdentifier", "Value": DBInstanceIdentifier}],
                },
                "Period": 86400 * 7,
                "Stat": "Minimum",
            },
        }],
        StartTime=(datetime.now() - timedelta(seconds=86400 * 7)).timestamp(),
        EndTime=datetime.now().timestamp(),
        ScanBy="TimestampDescending",
    )
    try:
        return round(response["MetricDataResults"][0]["Values"][0] / 1024**3, 2)
    except (IndexError, KeyError):
        return 0.0

def conditional(db_instance):
    return db_instance.get("DBInstanceClass", "") != "db.serverless" and db_instance.get("DBClusterIdentifier", "") == ""

def format_data_output(db_instances):
    return [
        (
            db_instance["DBInstanceIdentifier"],
            db_instance["DBInstanceClass"],
            db_instance["Engine"],
            db_instance["EngineVersion"],
            db_instance["AllocatedStorage"],
            get_db_freestorage(db_instance["DBInstanceIdentifier"]),
        )
        for db_instance in db_instances
    ]

def get_rds_instances():
    return [db_instance for db_instance in rds.describe_db_instances()["DBInstances"] if conditional(db_instance)]

def run():
    args = get_args()
    consolidated_rds_formatted_data = format_data_output(get_rds_instances())
    save_to_excel(consolidated_rds_formatted_data, ["DBInstanceIdentifier", "DBInstanceClass", "DBEngine", "Version", "AllocatedStorage", "FreeStorageSpace"], args.output)

if __name__ == "__main__":
    # run()
    print(format_data_output(get_rds_instances()))